from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.db.models import Sum, F, Count
from django.db.models.functions import TruncDate
from datetime import date
from inventory.models import Product, Category, Supplier
from sales.models import Sale
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from django.http import HttpResponse


@login_required
def dashboard(request):
    total_products = Product.objects.count()
    low_stock_items = Product.objects.filter(stock_quantity__lte=F('reorder_level'))
    all_products = Product.objects.select_related('category', 'supplier').all()
    categories = Category.objects.all()
    suppliers = Supplier.objects.all()

    today = date.today()
    today_sales = Sale.objects.filter(date__date=today)
    today_revenue = today_sales.aggregate(total=Sum('total_price'))['total'] or 0
    today_items_sold = today_sales.aggregate(total=Sum('quantity_sold'))['total'] or 0
    total_sales_value = Sale.objects.aggregate(total=Sum('total_price'))['total'] or 0
    recent_sales = Sale.objects.order_by('-date')[:10]

    daily_breakdown = (
        Sale.objects.annotate(day=TruncDate('date'))
        .values('day')
        .annotate(
            revenue=Sum('total_price'),
            items_sold=Sum('quantity_sold'),
            transactions=Count('id')
        )
        .order_by('-day')[:14]
    )

    chart_data_qs = list(reversed(list(daily_breakdown)))
    chart_labels = [str(entry['day']) for entry in chart_data_qs]
    chart_data = [float(entry['revenue']) for entry in chart_data_qs]

    context = {
        'total_products': total_products,
        'low_stock_items': low_stock_items,
        'all_products': all_products,
        'categories': categories,
        'suppliers': suppliers,
        'today_revenue': today_revenue,
        'today_items_sold': today_items_sold,
        'total_sales_value': total_sales_value,
        'recent_sales': recent_sales,
        'daily_breakdown': daily_breakdown,
        'chart_labels': chart_labels,
        'chart_data': chart_data,
        'today': today,
    }
    return render(request, 'reports/dashboard.html', context)


@login_required
def search_products(request):
    query = request.GET.get('q', '')
    results = Product.objects.filter(
        name__icontains=query
    ) if query else Product.objects.none()
    return render(request, 'reports/search.html', {
        'query': query,
        'results': results,
    })


@login_required
def restock_product(request, product_id):
    product = get_object_or_404(Product, id=product_id)
    if request.method == 'POST':
        quantity = int(request.POST.get('quantity', 0))
        if quantity > 0:
            product.stock_quantity += quantity
            product.save()
            messages.success(request, f'{product.name} restocked by {quantity} units!')
        return redirect('dashboard')
    return render(request, 'reports/restock.html', {'product': product})


@login_required
def add_product(request):
    categories = Category.objects.all()
    suppliers = Supplier.objects.all()

    if request.method == 'POST':
        form_type = request.POST.get('form_type')

        if form_type == 'add_product':
            name = request.POST.get('name')
            category_id = request.POST.get('category')
            supplier_id = request.POST.get('supplier')
            price = request.POST.get('price')
            stock_quantity = request.POST.get('stock_quantity')
            reorder_level = request.POST.get('reorder_level', 5)
            Product.objects.create(
                name=name,
                category_id=category_id if category_id else None,
                supplier_id=supplier_id if supplier_id else None,
                price=price,
                stock_quantity=stock_quantity,
                reorder_level=reorder_level,
            )
            messages.success(request, f'Product "{name}" added successfully!')

        elif form_type == 'add_category':
            name = request.POST.get('name')
            if name:
                Category.objects.create(name=name)
                messages.success(request, f'Category "{name}" added!')
            else:
                messages.error(request, 'Category name cannot be empty!')

        elif form_type == 'add_supplier':
            name = request.POST.get('name')
            if name:
                Supplier.objects.create(
                    name=name,
                    phone=request.POST.get('phone', ''),
                    email=request.POST.get('email', ''),
                )
                messages.success(request, f'Supplier "{name}" added!')
            else:
                messages.error(request, 'Supplier name cannot be empty!')

        return redirect('add_product')

    return render(request, 'reports/add_product.html', {
        'categories': categories,
        'suppliers': suppliers,
    })


@login_required
def edit_product(request, product_id):
    product = get_object_or_404(Product, id=product_id)
    categories = Category.objects.all()
    suppliers = Supplier.objects.all()
    if request.method == 'POST':
        product.name = request.POST.get('name')
        product.category_id = request.POST.get('category')
        product.supplier_id = request.POST.get('supplier')
        product.price = request.POST.get('price')
        product.stock_quantity = request.POST.get('stock_quantity')
        product.reorder_level = request.POST.get('reorder_level')
        product.save()
        messages.success(request, f'"{product.name}" updated!')
        return redirect('dashboard')
    return render(request, 'reports/edit_product.html', {
        'product': product, 'categories': categories, 'suppliers': suppliers
    })


@login_required
def delete_product(request, product_id):
    product = get_object_or_404(Product, id=product_id)
    if request.method == 'POST':
        name = product.name
        product.delete()
        messages.success(request, f'"{name}" deleted!')
    return redirect('dashboard')


@login_required
def delete_category(request, category_id):
    cat = get_object_or_404(Category, id=category_id)
    if request.method == 'POST':
        cat.delete()
        messages.success(request, 'Category deleted!')
    return redirect('add_product')


@login_required
def delete_supplier(request, supplier_id):
    sup = get_object_or_404(Supplier, id=supplier_id)
    if request.method == 'POST':
        sup.delete()
        messages.success(request, 'Supplier deleted!')
    return redirect('add_product')


@login_required
def record_sale(request):
    products = Product.objects.filter(stock_quantity__gt=0)

    # Initialize cart in session
    if 'cart' not in request.session:
        request.session['cart'] = []

    cart = request.session['cart']

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'add_to_cart':
            product_id = request.POST.get('product')
            quantity = int(request.POST.get('quantity', 1))
            product = get_object_or_404(Product, id=product_id)

            if quantity <= 0:
                messages.error(request, 'Quantity must be greater than 0!')
            elif quantity > product.stock_quantity:
                messages.error(request, f'Only {product.stock_quantity} units of {product.name} available!')
            else:
                # Check if product already in cart
                found = False
                for item in cart:
                    if item['product_id'] == int(product_id):
                        new_qty = item['quantity'] + quantity
                        if new_qty > product.stock_quantity:
                            messages.error(request, f'Total quantity exceeds available stock!')
                        else:
                            item['quantity'] = new_qty
                            item['total'] = float(product.price) * new_qty
                            messages.success(request, f'Updated {product.name} quantity in cart!')
                        found = True
                        break

                if not found:
                    cart.append({
                        'product_id': product.id,
                        'product_name': product.name,
                        'category': str(product.category),
                        'price': float(product.price),
                        'quantity': quantity,
                        'total': float(product.price) * quantity,
                    })
                    messages.success(request, f'{product.name} added to cart!')

                request.session['cart'] = cart
                request.session.modified = True

        elif action == 'remove_from_cart':
            product_id = int(request.POST.get('product_id'))
            cart = [item for item in cart if item['product_id'] != product_id]
            request.session['cart'] = cart
            request.session.modified = True
            messages.success(request, 'Item removed from cart!')

        elif action == 'clear_cart':
            request.session['cart'] = []
            request.session.modified = True
            messages.success(request, 'Cart cleared!')

        elif action == 'checkout':
            if not cart:
                messages.error(request, 'Cart is empty!')
            else:
                # Create receipt
                receipt_obj = Receipt.objects.create(
                    served_by=request.user.username,
                    total_amount=sum(item['total'] for item in cart)
                )

                # Create sale for each item
                for item in cart:
                    product = get_object_or_404(Product, id=item['product_id'])
                    Sale.objects.create(
                        receipt=receipt_obj,
                        receipt_number=receipt_obj.receipt_number,
                        product=product,
                        quantity_sold=item['quantity'],
                    )

                # Clear cart
                request.session['cart'] = []
                request.session.modified = True

                return redirect('receipt', receipt_id=receipt_obj.id)

        return redirect('record_sale')

    cart_total = sum(item['total'] for item in cart)
    cart_count = sum(item['quantity'] for item in cart)

    return render(request, 'reports/record_sale.html', {
        'products': products,
        'cart': cart,
        'cart_total': cart_total,
        'cart_count': cart_count,
    })


@login_required
def receipt(request, receipt_id):
    receipt_obj = get_object_or_404(Receipt, id=receipt_id)
    items = receipt_obj.items.select_related('product').all()
    return render(request, 'reports/receipt.html', {
        'receipt': receipt_obj,
        'items': items,
    })
@login_required
def filter_sales(request):
    sales = Sale.objects.all().order_by('-date')
    products = Product.objects.all()

    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    product_id = request.GET.get('product')

    if date_from:
        sales = sales.filter(date__date__gte=date_from)
    if date_to:
        sales = sales.filter(date__date__lte=date_to)
    if product_id:
        sales = sales.filter(product_id=product_id)

    total = sales.aggregate(total=Sum('total_price'))['total'] or 0
    items = sales.aggregate(total=Sum('quantity_sold'))['total'] or 0

    return render(request, 'reports/filter_sales.html', {
        'sales': sales,
        'products': products,
        'total': total,
        'items': items,
        'date_from': date_from or '',
        'date_to': date_to or '',
        'selected_product': product_id or '',
    })


@login_required
def print_report(request):
    today = date.today()
    date_from = request.GET.get('date_from', str(today))
    date_to = request.GET.get('date_to', str(today))

    if '-' not in date_from:
        date_from = str(today)
    if '-' not in date_to:
        date_to = str(today)

    sales = Sale.objects.filter(
        date__date__gte=date_from,
        date__date__lte=date_to
    ).order_by('date')

    total_revenue = sales.aggregate(total=Sum('total_price'))['total'] or 0
    total_items = sales.aggregate(total=Sum('quantity_sold'))['total'] or 0

    daily = (
        sales.annotate(day=TruncDate('date'))
        .values('day')
        .annotate(revenue=Sum('total_price'), items=Sum('quantity_sold'))
        .order_by('day')
    )

    return render(request, 'reports/print_report.html', {
        'sales': sales,
        'total_revenue': total_revenue,
        'total_items': total_items,
        'daily': daily,
        'date_from': date_from,
        'date_to': date_to,
    })


@login_required
def settings_page(request):
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'change_password':
            from django.contrib.auth import update_session_auth_hash
            old_password = request.POST.get('old_password')
            new_password = request.POST.get('new_password')
            confirm_password = request.POST.get('confirm_password')

            if not request.user.check_password(old_password):
                messages.error(request, 'Current password is incorrect.')
            elif new_password != confirm_password:
                messages.error(request, 'New passwords do not match.')
            elif len(new_password) < 6:
                messages.error(request, 'Password must be at least 6 characters.')
            else:
                request.user.set_password(new_password)
                request.user.save()
                update_session_auth_hash(request, request.user)
                messages.success(request, 'Password changed successfully!')

    return render(request, 'reports/settings.html')


@login_required
def export_excel(request):
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')

    sales = Sale.objects.all().order_by('-date')
    if date_from:
        sales = sales.filter(date__date__gte=date_from)
    if date_to:
        sales = sales.filter(date__date__lte=date_to)

    wb = openpyxl.Workbook()

    # Sheet 1 — Transactions
    ws1 = wb.active
    ws1.title = "Sales Transactions"
    header_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    center = Alignment(horizontal="center")

    ws1.merge_cells('A1:G1')
    ws1['A1'] = 'Hardware Store — Sales Report'
    ws1['A1'].font = Font(bold=True, size=14)
    ws1['A1'].alignment = center

    ws1.merge_cells('A2:G2')
    ws1['A2'] = f"Period: {date_from or 'All time'} to {date_to or 'Today'}"
    ws1['A2'].alignment = center
    ws1['A2'].font = Font(size=10, color="888888")

    headers = ['Receipt No', 'Product', 'Category', 'Unit Price (KSh)', 'Quantity', 'Total (KSh)', 'Date']
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=4, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    total_revenue = 0
    total_items = 0
    for row, sale in enumerate(sales, 5):
        ws1.cell(row=row, column=1, value=sale.receipt_number)
        ws1.cell(row=row, column=2, value=sale.product.name)
        ws1.cell(row=row, column=3, value=str(sale.product.category))
        ws1.cell(row=row, column=4, value=float(sale.product.price))
        ws1.cell(row=row, column=5, value=sale.quantity_sold)
        ws1.cell(row=row, column=6, value=float(sale.total_price))
        ws1.cell(row=row, column=7, value=sale.date.strftime('%d %b %Y, %H:%M'))
        if row % 2 == 0:
            for col in range(1, 8):
                ws1.cell(row=row, column=col).fill = PatternFill(
                    start_color="F0F2F5", end_color="F0F2F5", fill_type="solid"
                )
        total_revenue += float(sale.total_price)
        total_items += sale.quantity_sold

    total_row = len(list(sales)) + 5
    ws1.cell(row=total_row, column=1, value='TOTAL').font = Font(bold=True)
    ws1.cell(row=total_row, column=5, value=total_items).font = Font(bold=True)
    ws1.cell(row=total_row, column=6, value=total_revenue).font = Font(bold=True)
    for col in range(1, 8):
        ws1.cell(row=total_row, column=col).fill = PatternFill(
            start_color="D4EDDA", end_color="D4EDDA", fill_type="solid"
        )

    ws1.column_dimensions['A'].width = 16
    ws1.column_dimensions['B'].width = 22
    ws1.column_dimensions['C'].width = 16
    ws1.column_dimensions['D'].width = 18
    ws1.column_dimensions['E'].width = 12
    ws1.column_dimensions['F'].width = 16
    ws1.column_dimensions['G'].width = 22

    # Sheet 2 — Daily Summary
    ws2 = wb.create_sheet(title="Daily Summary")
    ws2.merge_cells('A1:D1')
    ws2['A1'] = 'Daily Sales Summary'
    ws2['A1'].font = Font(bold=True, size=14)
    ws2['A1'].alignment = center

    for col, header in enumerate(['Date', 'Transactions', 'Items Sold', 'Revenue (KSh)'], 1):
        cell = ws2.cell(row=3, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    daily = (
        sales.annotate(day=TruncDate('date'))
        .values('day')
        .annotate(
            revenue=Sum('total_price'),
            items=Sum('quantity_sold'),
            transactions=Count('id')
        )
        .order_by('-day')
    )

    for row, day in enumerate(daily, 4):
        ws2.cell(row=row, column=1, value=str(day['day']))
        ws2.cell(row=row, column=2, value=day['transactions'])
        ws2.cell(row=row, column=3, value=day['items'])
        ws2.cell(row=row, column=4, value=float(day['revenue']))
        if row % 2 == 0:
            for col in range(1, 5):
                ws2.cell(row=row, column=col).fill = PatternFill(
                    start_color="F0F2F5", end_color="F0F2F5", fill_type="solid"
                )

    ws2.column_dimensions['A'].width = 16
    ws2.column_dimensions['B'].width = 16
    ws2.column_dimensions['C'].width = 14
    ws2.column_dimensions['D'].width = 18

    # Sheet 3 — Product Summary
    ws3 = wb.create_sheet(title="Product Summary")
    ws3.merge_cells('A1:D1')
    ws3['A1'] = 'Sales by Product'
    ws3['A1'].font = Font(bold=True, size=14)
    ws3['A1'].alignment = center

    for col, header in enumerate(['Product', 'Category', 'Units Sold', 'Revenue (KSh)'], 1):
        cell = ws3.cell(row=3, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    product_summary = (
        sales.values('product__name', 'product__category__name')
        .annotate(units=Sum('quantity_sold'), revenue=Sum('total_price'))
        .order_by('-revenue')
    )

    for row, item in enumerate(product_summary, 4):
        ws3.cell(row=row, column=1, value=item['product__name'])
        ws3.cell(row=row, column=2, value=item['product__category__name'] or 'N/A')
        ws3.cell(row=row, column=3, value=item['units'])
        ws3.cell(row=row, column=4, value=float(item['revenue']))
        if row % 2 == 0:
            for col in range(1, 5):
                ws3.cell(row=row, column=col).fill = PatternFill(
                    start_color="F0F2F5", end_color="F0F2F5", fill_type="solid"
                )

    ws3.column_dimensions['A'].width = 24
    ws3.column_dimensions['B'].width = 16
    ws3.column_dimensions['C'].width = 14
    ws3.column_dimensions['D'].width = 18

    filename = f"hardware_store_sales_{date.today()}.xlsx"
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response